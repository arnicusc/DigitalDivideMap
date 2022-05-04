import geopandas as gp      #GIS data
import pandas as pd         #Dataframes
import argparse
from utils import download_from_drive #Download file from Drive

from os import path
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Year')
parser.add_argument('--year', default='2019',help="Year for which script should run")
args = parser.parse_known_args()[0]
year = args.year

bc_collapsed = f"../Data Outputs/CPUC/BC_Collapsed_{year}.csv" #Block Code Collapsed Data

print('Reading Data...')
bc_collapsed = pd.read_csv(bc_collapsed)

bc_collapsed['County'] = bc_collapsed['BlockCode'].astype('str').str[:4] #County
bc_collapsed['CountyPop'] = bc_collapsed['County'].map(bc_collapsed.groupby('County')['POP10'].sum().to_dict())
bc_collapsed['County_shrpop10'] = bc_collapsed['POP10'] / bc_collapsed['CountyPop']

#Temprory Columns for Population Weight
bc_collapsed['spop'] = bc_collapsed['served']*bc_collapsed['County_shrpop10']
bc_collapsed['spop25'] = bc_collapsed['served25']*bc_collapsed['County_shrpop10']
bc_collapsed['spop100'] = bc_collapsed['served100']*bc_collapsed['County_shrpop10']
bc_collapsed['spopfib'] = bc_collapsed['served_fib']*bc_collapsed['County_shrpop10']
bc_collapsed['cpop'] = bc_collapsed['comp']*bc_collapsed['County_shrpop10']
bc_collapsed['cpop25'] = bc_collapsed['comp25']*bc_collapsed['County_shrpop10']
bc_collapsed['cpop100'] = bc_collapsed['comp100']*bc_collapsed['County_shrpop10']
bc_collapsed['cpopfib'] = bc_collapsed['comp_fib']*bc_collapsed['County_shrpop10']

bc_collapsed['avgup'] = bc_collapsed['max_up']*bc_collapsed['County_shrpop10']
bc_collapsed['avgdn'] = bc_collapsed['max_dn']*bc_collapsed['County_shrpop10']

print("Performing transformations...")
#Dataframe collapsed by Block Group
cn_collapsed = pd.DataFrame()
cn_collapsed['County'] = sorted(bc_collapsed['BlockCode'].astype('str').str[:4].unique()) #County
cn_collapsed['County'] = '0' + cn_collapsed['County'].astype('str')
cn_collapsed['CountyPop'] = bc_collapsed.groupby('County').CountyPop.mean().values
cn_collapsed['LACounty'] = bc_collapsed.groupby('County').LACounty.mean().values #LACounty or not

#Served and not served weighed by population
cn_collapsed['served'] = bc_collapsed.groupby('County').spop.sum().values*100
cn_collapsed['served25'] = bc_collapsed.groupby('County').spop25.sum().values*100
cn_collapsed['served100'] = bc_collapsed.groupby('County').spop100.sum().values*100
cn_collapsed['served_fib'] = bc_collapsed.groupby('County').spopfib.sum().values*100
# cn_collapsed['not_served'] = 1.0 - cn_collapsed['served']
# cn_collapsed['served_under25'] = cn_collapsed['served'] - cn_collapsed['served25']
# cn_collapsed['served_under100'] = cn_collapsed['served'] - cn_collapsed['served100']

#Comp and no comp weighed by population
cn_collapsed['comp'] = bc_collapsed.groupby('County').cpop.sum().values*100
cn_collapsed['comp25'] = bc_collapsed.groupby('County').cpop25.sum().values*100
cn_collapsed['comp100'] = bc_collapsed.groupby('County').cpop100.sum().values*100
cn_collapsed['comp_fib'] = bc_collapsed.groupby('County').cpopfib.sum().values*100
# cn_collapsed['no_comp'] = 1.0 - cn_collapsed['comp']
# cn_collapsed['comp_under25'] = cn_collapsed['comp'] - cn_collapsed['comp25']
# cn_collapsed['comp_under100'] = cn_collapsed['comp'] - cn_collapsed['comp100']

#Max upload and download speed
# cn_collapsed['max_up'] = bc_collapsed.groupby('County').max_up.max().values
# cn_collapsed['max_dn'] = bc_collapsed.groupby('County').max_dn.max().values

#Ask Doubt
#Average max upload and download weighed by population 
cn_collapsed['avg_max_up'] = bc_collapsed.groupby('County').avgup.sum().values
cn_collapsed['avg_max_dn'] = bc_collapsed.groupby('County').avgdn.sum().values

cn_collapsed = cn_collapsed.round(1)

print("Saving File...")
filepath = r"../Data Outputs/CPUC/County_Collapsed.xlsx"

if path.exists(filepath) == False:
    pd.DataFrame().to_excel(filepath)

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine = 'openpyxl', if_sheet_exists="error")
writer.book = book

cn_collapsed.to_excel(writer, sheet_name = f"county_{year}", index=False) #BG_Collapsed_2019

if "Sheet1" in writer.book.sheetnames:
    std=writer.book['Sheet1']
    writer.book.remove(std)

writer.save()

print(f"Finished! File saved to ../Data Outputs/CPUC/CN_Collapsed.xlsx")
