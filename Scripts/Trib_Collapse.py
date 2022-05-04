import geopandas as gp      #GIS data
import pandas as pd         #Dataframes
import argparse

from os import path
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Year')
parser.add_argument('--year', default='2019',help="Year for which script should run")
args = parser.parse_known_args()[0]
year = args.year

tri = gp.read_file("zip://../Data Sources/BIA_National_LAR_shp.zip!BIA_National_LAR.shp")
if int(year) >= 2020:
    bc = gp.read_file("../Data Sources/Census/tl_2020_06_tabblock20.zip")
else:
    bc = gp.read_file("../Data Sources/Census/tl_2010_06_tabblock10.zip")

bc_tri = gp.sjoin(bc,tri,how = 'inner', op = 'intersects')

bc_data = pd.read_csv(f"../Data Outputs/CPUC/BC_Collapsed_{year}.csv")
bc_data['BlockCode'] = '0' + bc_data['BlockCode'].astype('str')

if int(year) >=2020:
    bc_tri = bc_tri.merge(bc_data, left_on='GEOID20', right_on='BlockCode')
else:
    bc_tri = bc_tri.merge(bc_data, left_on='GEOID10', right_on='BlockCode')



bc_tri['trib_pop'] = bc_tri['LARID'].map(bc_tri.groupby('LARID').POP10.sum().to_dict())
bc_tri['trib_shrpop10'] = bc_tri['POP10'] / bc_tri['trib_pop']

bc_tri['spop'] = bc_tri['served']*bc_tri['trib_shrpop10']
bc_tri['spop25'] = bc_tri['served25']*bc_tri['trib_shrpop10']
bc_tri['spop100'] = bc_tri['served100']*bc_tri['trib_shrpop10']
bc_tri['spopfib'] = bc_tri['served_fib']*bc_tri['trib_shrpop10']
bc_tri['cpop'] = bc_tri['comp']*bc_tri['trib_shrpop10']
bc_tri['cpop25'] = bc_tri['comp25']*bc_tri['trib_shrpop10']
bc_tri['cpop100'] = bc_tri['comp100']*bc_tri['trib_shrpop10']
bc_tri['cpopfib'] = bc_tri['comp_fib']*bc_tri['trib_shrpop10']

bc_tri['avgup'] = bc_tri['max_up']*bc_tri['trib_shrpop10']
bc_tri['avgdn'] = bc_tri['max_dn']*bc_tri['trib_shrpop10']

tri_collapsed = pd.DataFrame()
tri_collapsed['LARID'] = sorted(bc_tri['LARID'].unique())
tri_collapsed['trib_pop'] = bc_tri.groupby('LARID').trib_pop.mean().values

tri_collapsed['served'] = bc_tri.groupby('LARID').spop.sum().values*100
tri_collapsed['served25'] = bc_tri.groupby('LARID').spop25.sum().values*100
tri_collapsed['served100'] = bc_tri.groupby('LARID').spop100.sum().values*100
tri_collapsed['served_fib'] = bc_tri.groupby('LARID').spopfib.sum().values*100
# tri_collapsed['not_served'] = 1.0 - tri_collapsed['served']
# tri_collapsed['served_under25'] = tri_collapsed['served'] - tri_collapsed['served25']
# tri_collapsed['served_under100'] = tri_collapsed['served'] - tri_collapsed['served100']

#Comp and no comp weighed by population
tri_collapsed['comp'] = bc_tri.groupby('LARID').cpop.sum().values*100
tri_collapsed['comp25'] = bc_tri.groupby('LARID').cpop25.sum().values*100
tri_collapsed['comp100'] = bc_tri.groupby('LARID').cpop100.sum().values*100
tri_collapsed['comp_fib'] = bc_tri.groupby('LARID').cpopfib.sum().values*100
# tri_collapsed['no_comp'] = 1.0 - tri_collapsed['comp']
# tri_collapsed['comp_under25'] = tri_collapsed['comp'] - tri_collapsed['comp25']
# tri_collapsed['comp_under100'] = tri_collapsed['comp'] - tri_collapsed['comp100']

#Max upload and download speed
# tri_collapsed['max_up'] = bc_tri.groupby('LARID').max_up.max().values
# tri_collapsed['max_dn'] = bc_tri.groupby('LARID').max_dn.max().values

#Ask Doubt
#Average max upload and download weighed by population 
tri_collapsed['avg_max_up'] = bc_tri.groupby('LARID').avgup.sum().values
tri_collapsed['avg_max_dn'] = bc_tri.groupby('LARID').avgdn.sum().values

tri_collapsed = tri_collapsed.round(1)

print("Saving File...")
filepath = r"../Data Outputs/CPUC/Tri_Collapsed.xlsx"

if path.exists(filepath) == False:
    pd.DataFrame().to_excel(filepath)

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine = 'openpyxl', if_sheet_exists="error")
writer.book = book

tri_collapsed.to_excel(writer, sheet_name = f"tri_{year}", index=False) #BG_Collapsed_2019

if "Sheet1" in writer.book.sheetnames:
    std=writer.book['Sheet1']
    writer.book.remove(std)

writer.save()

print(f"Finished! File saved to ../Data Outputs/CPUC/Tri_Collapsed.xlsx")
