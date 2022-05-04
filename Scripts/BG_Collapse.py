import geopandas as gp      #GIS data
import pandas as pd         #Dataframes
import argparse
from os import path
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Year')
parser.add_argument('--year', default='2019',help="Year for which script should run")
args = parser.parse_known_args()[0]
year = args.year

bc_collapsed = f"../Data Outputs/CPUC/BC_Collapsed_{year}.csv" #Block Code Collapsed Data

print('Reading Data...')
bc_collapsed = pd.read_csv(bc_collapsed)

#Temprory Columns for Population Weight
bc_collapsed['spop'] = bc_collapsed['served']*bc_collapsed['shrpop10']
bc_collapsed['spop25'] = bc_collapsed['served25']*bc_collapsed['shrpop10']
bc_collapsed['spop100'] = bc_collapsed['served100']*bc_collapsed['shrpop10']
bc_collapsed['spopfib'] = bc_collapsed['served_fib']*bc_collapsed['shrpop10']
bc_collapsed['cpop'] = bc_collapsed['comp']*bc_collapsed['shrpop10']
bc_collapsed['cpop25'] = bc_collapsed['comp25']*bc_collapsed['shrpop10']
bc_collapsed['cpop100'] = bc_collapsed['comp100']*bc_collapsed['shrpop10']
bc_collapsed['cpopfib'] = bc_collapsed['comp_fib']*bc_collapsed['shrpop10']

bc_collapsed['avgup'] = bc_collapsed['max_up']*bc_collapsed['shrpop10']
bc_collapsed['avgdn'] = bc_collapsed['max_dn']*bc_collapsed['shrpop10']

print("Performing transformations...")
#Dataframe collapsed by Block Group
bg_collapsed = pd.DataFrame()
bg_collapsed['BlockGroup'] = sorted(bc_collapsed['BlockGroup'].unique()) #Block Group
bg_collapsed['BlockGroup'] = '0' + bg_collapsed['BlockGroup'].astype('str')
bg_collapsed['BGPop'] = bc_collapsed.groupby('BlockGroup').BGPop.mean().values #Block Group Population
bg_collapsed['LACounty'] = bc_collapsed.groupby('BlockGroup').LACounty.mean().values #LACounty or not

#Served and not served weighed by population
bg_collapsed['served'] = bc_collapsed.groupby('BlockGroup').spop.sum().values*100
bg_collapsed['served25'] = bc_collapsed.groupby('BlockGroup').spop25.sum().values*100
bg_collapsed['served100'] = bc_collapsed.groupby('BlockGroup').spop100.sum().values*100
bg_collapsed['served_fib'] = bc_collapsed.groupby('BlockGroup').spopfib.sum().values*100
# bg_collapsed['not_served'] = 1.0 - bg_collapsed['served']
# bg_collapsed['served_under25'] = bg_collapsed['served'] - bg_collapsed['served25']
# bg_collapsed['served_under100'] = bg_collapsed['served'] - bg_collapsed['served100']

#Comp and no comp weighed by population
bg_collapsed['comp'] = bc_collapsed.groupby('BlockGroup').cpop.sum().values*100
bg_collapsed['comp25'] = bc_collapsed.groupby('BlockGroup').cpop25.sum().values*100
bg_collapsed['comp100'] = bc_collapsed.groupby('BlockGroup').cpop100.sum().values*100
bg_collapsed['comp_fib'] = bc_collapsed.groupby('BlockGroup').cpopfib.sum().values*100
# bg_collapsed['no_comp'] = 1.0 - bg_collapsed['comp']
# bg_collapsed['comp_under25'] = bg_collapsed['comp'] - bg_collapsed['comp25']
# bg_collapsed['comp_under100'] = bg_collapsed['comp'] - bg_collapsed['comp100']

#Max upload and download speed
# bg_collapsed['max_up'] = bc_collapsed.groupby('BlockGroup').max_up.max().values
# bg_collapsed['max_dn'] = bc_collapsed.groupby('BlockGroup').max_dn.max().values

#Ask Doubt
#Average max upload and download weighed by population 
bg_collapsed['avg_max_up'] = bc_collapsed.groupby('BlockGroup').avgup.sum().values
bg_collapsed['avg_max_dn'] = bc_collapsed.groupby('BlockGroup').avgdn.sum().values

bg_collapsed = bg_collapsed.round(1)


print("Saving File...")
filepath = r"../Data Outputs/CPUC/BG_Collapsed.xlsx"

if path.exists(filepath) == False:
    pd.DataFrame().to_excel(filepath)

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine = 'openpyxl', if_sheet_exists="error")
writer.book = book

bg_collapsed.to_excel(writer,sheet_name=f"bg_{year}",index=False) #BG_Collapsed_2019

if "Sheet1" in writer.book.sheetnames:
    std=writer.book['Sheet1']
    writer.book.remove(std)

writer.save()

print(f"Finished! File saved to ../Data Outputs/CPUC/BG_Collapsed.xlsx")
