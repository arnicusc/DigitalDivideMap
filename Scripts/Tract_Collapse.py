import geopandas as gp      #GIS data
import pandas as pd         #Dataframes
import argparse

from os import path
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Year')
parser.add_argument('--year', default='2019',help="Year for which script should run")
args = parser.parse_known_args()[0]
year = args.year

bc_collapsed = f"../Data Outputs/CPUC/BC_Collapsed_{year}.csv" #Block Code Collapsed Data

print('Reading Data...')
bc_collapsed = pd.read_csv(bc_collapsed)


bc_collapsed['Tract'] = bc_collapsed['BlockCode'].astype('str').str[:10] #County
bc_collapsed['TractPop'] = bc_collapsed['Tract'].map(bc_collapsed.groupby('Tract')['POP10'].sum().to_dict())
bc_collapsed['Tract_shrpop10'] = bc_collapsed['POP10'] / bc_collapsed['TractPop']

#Temprory Columns for Population Weight
bc_collapsed['spop'] = bc_collapsed['served']*bc_collapsed['Tract_shrpop10']
bc_collapsed['spop25'] = bc_collapsed['served25']*bc_collapsed['Tract_shrpop10']
bc_collapsed['spop100'] = bc_collapsed['served100']*bc_collapsed['Tract_shrpop10']
bc_collapsed['spopfib'] = bc_collapsed['served_fib']*bc_collapsed['Tract_shrpop10']
bc_collapsed['cpop'] = bc_collapsed['comp']*bc_collapsed['Tract_shrpop10']
bc_collapsed['cpop25'] = bc_collapsed['comp25']*bc_collapsed['Tract_shrpop10']
bc_collapsed['cpop100'] = bc_collapsed['comp100']*bc_collapsed['Tract_shrpop10']
bc_collapsed['cpopfib'] = bc_collapsed['comp_fib']*bc_collapsed['Tract_shrpop10']

bc_collapsed['avgup'] = bc_collapsed['max_up']*bc_collapsed['Tract_shrpop10']
bc_collapsed['avgdn'] = bc_collapsed['max_dn']*bc_collapsed['Tract_shrpop10']

print("Performing transformations...")
#Dataframe collapsed by Block Group
tr_collapsed = pd.DataFrame()
tr_collapsed['Tract'] = sorted(bc_collapsed['BlockCode'].astype('str').str[:10].unique()) #County
tr_collapsed['Tract'] = '0' + tr_collapsed['Tract'].astype('str')
tr_collapsed['TractPop'] = bc_collapsed.groupby('Tract').TractPop.mean().values
tr_collapsed['LACounty'] = bc_collapsed.groupby('Tract').LACounty.mean().values #LACounty or not

#Served and not served weighed by population
tr_collapsed['served'] = bc_collapsed.groupby('Tract').spop.sum().values*100
tr_collapsed['served25'] = bc_collapsed.groupby('Tract').spop25.sum().values*100
tr_collapsed['served100'] = bc_collapsed.groupby('Tract').spop100.sum().values*100
tr_collapsed['served_fib'] = bc_collapsed.groupby('Tract').spopfib.sum().values*100
# tr_collapsed['not_served'] = 1.0 - tr_collapsed['served']
# tr_collapsed['served_under25'] = tr_collapsed['served'] - tr_collapsed['served25']
# tr_collapsed['served_under100'] = tr_collapsed['served'] - tr_collapsed['served100']

#Comp and no comp weighed by population
tr_collapsed['comp'] = bc_collapsed.groupby('Tract').cpop.sum().values*100
tr_collapsed['comp25'] = bc_collapsed.groupby('Tract').cpop25.sum().values*100
tr_collapsed['comp100'] = bc_collapsed.groupby('Tract').cpop100.sum().values*100
tr_collapsed['comp_fib'] = bc_collapsed.groupby('Tract').cpopfib.sum().values*100
# tr_collapsed['no_comp'] = 1.0 - tr_collapsed['comp']
# tr_collapsed['comp_under25'] = tr_collapsed['comp'] - tr_collapsed['comp25']
# tr_collapsed['comp_under100'] = tr_collapsed['comp'] - tr_collapsed['comp100']

#Max upload and download speed
# tr_collapsed['max_up'] = bc_collapsed.groupby('Tract').max_up.max().values
# tr_collapsed['max_dn'] = bc_collapsed.groupby('Tract').max_dn.max().values

#Ask Doubt
#Average max upload and download weighed by population 
tr_collapsed['avg_max_up'] = bc_collapsed.groupby('Tract').avgup.sum().values
tr_collapsed['avg_max_dn'] = bc_collapsed.groupby('Tract').avgdn.sum().values

tr_collapsed = tr_collapsed.round(1)

print("Saving File...")
filepath = r"../Data Outputs/CPUC/Tract_Collapsed.xlsx"

if path.exists(filepath) == False:
    pd.DataFrame().to_excel(filepath)

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine = 'openpyxl', if_sheet_exists="error")
writer.book = book

tr_collapsed.to_excel(writer, sheet_name = f"tr_{year}", index=False) #BG_Collapsed_2019

if "Sheet1" in writer.book.sheetnames:
    std=writer.book['Sheet1']
    writer.book.remove(std)

writer.save()

print(f"Finished! File saved to ../Data Outputs/CPUC/Tract_Collapsed.xlsx")
