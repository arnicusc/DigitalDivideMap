import geopandas as gp      #GIS data
import pandas as pd         #Dataframes
import argparse
from os import path
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Year')
parser.add_argument('--year', default='2019',help="Year for which script should run")
args = parser.parse_known_args()[0]
year = args.year

bc_collapsed = f"../Data Outputs/CPUC/BC_Collapsed_{year}.csv" #Block Code Collapsed Data

if int(year) >=2020:
    bg = gp.read_file('../Data Sources/Census/tl_2020_06_bg20.zip')
else:
    bg = gp.read_file('../Data Sources/Census/tl_2010_06_bg10.zip')

if int(year) >= 2020:
    cd = gp.read_file('../Data Sources/Census/tl_2020_06_cd113.zip')
else:
    cd = gp.read_file('../Data Sources/Census/tl_2010_06_cd111.zip')


cd_bg = gp.sjoin(cd,bg, how = 'inner', op = 'intersects')
if int(year) >= 2020:
    cd_bg = cd_bg.rename(columns={'GEOID10':'CD','GEOID20':'BlockGroup'})
else:
    cd_bg = cd_bg.rename(columns={'GEOID10_left':'CD','GEOID10_right':'BlockGroup'})
    
cd_bg = cd_bg[['CD','BlockGroup']]
cd_bg = cd_bg.loc[:,~cd_bg.columns.duplicated()]
cd_bg = cd_bg.set_index('BlockGroup').squeeze()

print('Reading Data...')
bc_collapsed = pd.read_csv(bc_collapsed)


bc_collapsed['BlockGroup'] = '0' + bc_collapsed['BlockGroup'].astype('str')
bc_collapsed['CD'] = bc_collapsed['BlockGroup'].map(cd_bg.to_dict())
bc_collapsed['CDPop'] = bc_collapsed['CD'].map(bc_collapsed.groupby('CD')['POP10'].sum().to_dict())
bc_collapsed['CD_shrpop10'] = bc_collapsed['POP10'] / bc_collapsed['CDPop']
                                                                   
#Temprory Columns for Population Weight
bc_collapsed['spop'] = bc_collapsed['served']*bc_collapsed['CD_shrpop10']
bc_collapsed['spop25'] = bc_collapsed['served25']*bc_collapsed['CD_shrpop10']
bc_collapsed['spop100'] = bc_collapsed['served100']*bc_collapsed['CD_shrpop10']
bc_collapsed['spopfib'] = bc_collapsed['served_fib']*bc_collapsed['CD_shrpop10']
bc_collapsed['cpop'] = bc_collapsed['comp']*bc_collapsed['CD_shrpop10']
bc_collapsed['cpop25'] = bc_collapsed['comp25']*bc_collapsed['CD_shrpop10']
bc_collapsed['cpop100'] = bc_collapsed['comp100']*bc_collapsed['CD_shrpop10']
bc_collapsed['cpopfib'] = bc_collapsed['comp_fib']*bc_collapsed['CD_shrpop10']

bc_collapsed['avgup'] = bc_collapsed['max_up']*bc_collapsed['CD_shrpop10']
bc_collapsed['avgdn'] = bc_collapsed['max_dn']*bc_collapsed['CD_shrpop10']

print("Performing transformations...")
#Dataframe collapsed by Block Group
cd_collapsed = pd.DataFrame()
cd_collapsed['CD'] = sorted(bc_collapsed['CD'].unique()) #County
# cd_collapsed['CD'] = '0' + cd_collapsed['CD'].astype('str')
cd_collapsed['CDPop'] = bc_collapsed.groupby('CD').CDPop.mean().values
# cd_collapsed['LACounty'] = bc_collapsed.groupby('County').LACounty.mean().values #LACounty or not

#Served and not served weighed by population
cd_collapsed['served'] = bc_collapsed.groupby('CD').spop.sum().values*100
cd_collapsed['served25'] = bc_collapsed.groupby('CD').spop25.sum().values*100
cd_collapsed['served100'] = bc_collapsed.groupby('CD').spop100.sum().values*100
cd_collapsed['served_fib'] = bc_collapsed.groupby('CD').spopfib.sum().values*100
# cd_collapsed['not_served'] = 1.0 - cd_collapsed['served']
# cd_collapsed['served_under25'] = cd_collapsed['served'] - cd_collapsed['served25']
# cd_collapsed['served_under100'] = cd_collapsed['served'] - cd_collapsed['served100']

#Comp and no comp weighed by population
cd_collapsed['comp'] = bc_collapsed.groupby('CD').cpop.sum().values*100
cd_collapsed['comp25'] = bc_collapsed.groupby('CD').cpop25.sum().values*100
cd_collapsed['comp100'] = bc_collapsed.groupby('CD').cpop100.sum().values*100
cd_collapsed['comp_fib'] = bc_collapsed.groupby('CD').cpopfib.sum().values*100
# cd_collapsed['no_comp'] = 1.0 - cd_collapsed['comp']
# cd_collapsed['comp_under25'] = cd_collapsed['comp'] - cd_collapsed['comp25']
# cd_collapsed['comp_under100'] = cd_collapsed['comp'] - cd_collapsed['comp100']

#Max upload and download speed
# cd_collapsed['max_up'] = bc_collapsed.groupby('CD').max_up.max().values
# cd_collapsed['max_dn'] = bc_collapsed.groupby('CD').max_dn.max().values

#Ask Doubt
#Average max upload and download weighed by population 
cd_collapsed['avg_max_up'] = bc_collapsed.groupby('CD').avgup.sum().values
cd_collapsed['avg_max_dn'] = bc_collapsed.groupby('CD').avgdn.sum().values

cd_collapsed = cd_collapsed.round(1)


print("Saving File...")
filepath = r"../Data Outputs/CPUC/CD_Collapsed.xlsx"

if path.exists(filepath) == False:
    pd.DataFrame().to_excel(filepath)

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine = 'openpyxl', if_sheet_exists="error")
writer.book = book

cd_collapsed.to_excel(writer, sheet_name = f"cd_{year}", index=False) #BG_Collapsed_2019

if "Sheet1" in writer.book.sheetnames:
    std=writer.book['Sheet1']
    writer.book.remove(std)

writer.save()

print(f"Finished! File saved to ../Data Outputs/CPUC/CD_Collapsed.xlsx")
