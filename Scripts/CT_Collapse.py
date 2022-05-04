import geopandas as gp      #GIS data
import pandas as pd         #Dataframes
import argparse

from os import path
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Year')
parser.add_argument('--year', default='2019',help="Year for which script should run")
args = parser.parse_known_args()[0]
year = args.year


if int(year) >=2020:
    bc = gp.read_file('../Data Sources/Census/tl_2020_06_tabblock20.zip')
    ct = gp.read_file("../Data Sources/city_boundaries20.zip")
else:
    bc = gp.read_file('../Data Sources/Census/tl_2010_06_tabblock10.zip')
    ct = gp.read_file("../Data Sources/city_boundaries.zip")


bc_ct = gp.sjoin(bc,ct,how = 'inner', op = 'intersects')
bc_data = pd.read_csv(f"../Data Outputs/CPUC/BC_Collapsed_{year}.csv")
bc_data['BlockCode'] = '0' + bc_data['BlockCode'].astype('str')

if int(year) >=2020:
    bc_ct = bc_ct.merge(bc_data, left_on='GEOID20_left', right_on='BlockCode')
    bc_ct = bc_ct.rename(columns={'GEOID20_right': 'GEOID'})
else:
    bc_ct = bc_ct.merge(bc_data, left_on='GEOID10_left', right_on='BlockCode')
    bc_ct = bc_ct.rename(columns={'GEOID10_right': 'GEOID'})


bc_ct['city_pop'] = bc_ct['GEOID'].map(bc_ct.groupby('GEOID').POP10.sum().to_dict())
bc_ct['city_shrpop10'] = bc_ct['POP10'] / bc_ct['city_pop']

bc_ct['spop'] = bc_ct['served']*bc_ct['city_shrpop10']
bc_ct['spop25'] = bc_ct['served25']*bc_ct['city_shrpop10']
bc_ct['spop100'] = bc_ct['served100']*bc_ct['city_shrpop10']
bc_ct['spopfib'] = bc_ct['served_fib']*bc_ct['city_shrpop10']
bc_ct['cpop'] = bc_ct['comp']*bc_ct['city_shrpop10']
bc_ct['cpop25'] = bc_ct['comp25']*bc_ct['city_shrpop10']
bc_ct['cpop100'] = bc_ct['comp100']*bc_ct['city_shrpop10']
bc_ct['cpopfib'] = bc_ct['comp_fib']*bc_ct['city_shrpop10']

bc_ct['avgup'] = bc_ct['max_up']*bc_ct['city_shrpop10']
bc_ct['avgdn'] = bc_ct['max_dn']*bc_ct['city_shrpop10']

ct_collapsed = pd.DataFrame()
ct_collapsed['city'] = sorted(bc_ct['GEOID'].unique())
ct_collapsed['city_pop'] = bc_ct.groupby('GEOID').city_pop.mean().values

ct_collapsed['served'] = bc_ct.groupby('GEOID').spop.sum().values*100
ct_collapsed['served25'] = bc_ct.groupby('GEOID').spop25.sum().values*100
ct_collapsed['served100'] = bc_ct.groupby('GEOID').spop100.sum().values*100
ct_collapsed['served_fib'] = bc_ct.groupby('GEOID').spopfib.sum().values*100
# ct_collapsed['not_served'] = 1.0 - ct_collapsed['served']
# ct_collapsed['served_under25'] = ct_collapsed['served'] - ct_collapsed['served25']
# ct_collapsed['served_under100'] = ct_collapsed['served'] - ct_collapsed['served100']

#Comp and no comp weighed by population
ct_collapsed['comp'] = bc_ct.groupby('GEOID').cpop.sum().values*100
ct_collapsed['comp25'] = bc_ct.groupby('GEOID').cpop25.sum().values*100
ct_collapsed['comp100'] = bc_ct.groupby('GEOID').cpop100.sum().values*100
ct_collapsed['comp_fib'] = bc_ct.groupby('GEOID').cpopfib.sum().values*100
# ct_collapsed['no_comp'] = 1.0 - ct_collapsed['comp']
# ct_collapsed['comp_under25'] = ct_collapsed['comp'] - ct_collapsed['comp25']
# ct_collapsed['comp_under100'] = ct_collapsed['comp'] - ct_collapsed['comp100']

#Max upload and download speed
# ct_collapsed['max_up'] = bc_ct.groupby('GEOID').max_up.max().values
# ct_collapsed['max_dn'] = bc_ct.groupby('GEOID').max_dn.max().values

#Ask Doubt
#Average max upload and download weighed by population 
ct_collapsed['avg_max_up'] = bc_ct.groupby('GEOID').avgup.sum().values
ct_collapsed['avg_max_dn'] = bc_ct.groupby('GEOID').avgdn.sum().values

ct_collapsed = ct_collapsed.round(1)

print("Saving File...")
filepath = r"../Data Outputs/CPUC/CT_Collapsed.xlsx"

if path.exists(filepath) == False:
    pd.DataFrame().to_excel(filepath)

book = load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine = 'openpyxl', if_sheet_exists="error")
writer.book = book

ct_collapsed.to_excel(writer, sheet_name = f"ct_{year}", index=False) #BG_Collapsed_2019

if "Sheet1" in writer.book.sheetnames:
    std=writer.book['Sheet1']
    writer.book.remove(std)

writer.save()

print(f"Finished! File saved to ../Data Outputs/CPUC/CT_Collapsed.xlsx")

